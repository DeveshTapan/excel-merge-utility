#!/usr/bin/env python3
"""Merge Excel workbooks using their first worksheet and a strict schema."""

from __future__ import annotations

import argparse
import glob
import hashlib
import os
import queue
import secrets
import shutil
import subprocess
import sys
import tempfile
import threading
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Sequence

import numpy as np
import pandas as pd


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
EXCEL_MAX_ROWS = 1_048_576
INVALID_SHEET_CHARS = set(r"[]:*?/\\")
ProgressCallback = Callable[[int, int, str], None]


class MergeError(RuntimeError):
    """A user-facing merge error."""


class SchemaMismatchError(MergeError):
    """Raised when a workbook does not match the baseline schema."""


@dataclass(frozen=True)
class VerificationResult:
    passed: bool
    reason: str
    merged_rows: int
    expected_rows: int
    columns_equal: bool
    checksum_equal: bool | None = None
    checksum_merged: str | None = None
    checksum_expected: str | None = None
    diff_csv: Path | None = None
    per_column_mismatch_counts: dict[str, int] = field(default_factory=dict)


@dataclass(frozen=True)
class MergeResult:
    output_path: Path
    total_rows: int
    per_file_rows: tuple[tuple[Path, int], ...]
    verification: VerificationResult | None
    moved_files: tuple[tuple[Path, Path], ...] = ()
    move_errors: tuple[str, ...] = ()


def _stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")


def is_excel_file(path: os.PathLike[str] | str) -> bool:
    return Path(path).suffix.lower() in EXCEL_EXTENSIONS


def infer_engine(path: os.PathLike[str] | str) -> str:
    return "xlrd" if Path(path).suffix.lower() == ".xls" else "openpyxl"


def default_output_name(base_name: str = "merged") -> str:
    return f"{base_name}-{_stamp()}-{secrets.token_hex(3).upper()}.xlsx"


def sanitize_sheet_name(name: str) -> str:
    cleaned = "".join("_" if char in INVALID_SHEET_CHARS else char for char in name).strip().strip("'")
    if not cleaned:
        raise MergeError("The output sheet name cannot be empty.")
    return cleaned[:31]


def expand_patterns(patterns: Sequence[str], sort_mode: str) -> list[Path]:
    matches: list[Path] = []
    for pattern in patterns:
        found = [Path(item) for item in glob.glob(pattern)]
        if sort_mode == "name":
            found.sort(key=lambda item: str(item).casefold())
        matches.extend(found)
    return matches


def normalize_input_paths(paths: Iterable[os.PathLike[str] | str]) -> list[Path]:
    normalized: list[Path] = []
    seen: set[str] = set()
    for raw_path in paths:
        path = Path(raw_path).expanduser().resolve()
        key = os.path.normcase(str(path))
        if key in seen:
            continue
        seen.add(key)
        normalized.append(path)
    return normalized


def validate_input_paths(paths: Sequence[Path]) -> None:
    if not paths:
        raise MergeError("Select at least one Excel file.")
    for path in paths:
        if not path.exists():
            raise MergeError(f"Input file does not exist: {path.name}")
        if not path.is_file():
            raise MergeError(f"Input path is not a file: {path.name}")
        if not is_excel_file(path):
            raise MergeError(f"Unsupported Excel file type: {path.name}")
        if not os.access(path, os.R_OK):
            raise MergeError(f"Input file cannot be read: {path.name}")


def validate_columns(columns: Sequence[str], source_name: str) -> None:
    empty = [index + 1 for index, name in enumerate(columns) if not str(name).strip()]
    if empty:
        raise MergeError(f"{source_name} has blank column headings at positions: {empty}")
    duplicates = sorted({name for name in columns if columns.count(name) > 1})
    if duplicates:
        raise MergeError(f"{source_name} has duplicate column headings: {duplicates}")


def read_original_headers(path: Path) -> list[str]:
    try:
        if path.suffix.lower() == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(path, on_demand=True)
            try:
                values = workbook.sheet_by_index(0).row_values(0)
            finally:
                workbook.release_resources()
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                sheet = workbook.worksheets[0]
                values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            finally:
                workbook.close()
    except StopIteration as exc:
        raise MergeError(f"{path.name} does not contain a header row.") from exc
    except Exception as exc:
        raise MergeError(
            f"Could not inspect {path.name}. It may be damaged, password-protected, or unsupported."
        ) from exc
    headers = ["" if value is None else str(value) for value in values]
    validate_columns(headers, path.name)
    return headers


def read_first_sheet(path: Path, force_text: bool = True) -> pd.DataFrame:
    original_headers = read_original_headers(path)
    kwargs: dict[str, object] = {"engine": infer_engine(path), "sheet_name": 0}
    if force_text:
        kwargs["dtype"] = str
    try:
        frame = pd.read_excel(path, **kwargs)
    except PermissionError as exc:
        raise MergeError(f"Close the workbook and try again: {path.name}") from exc
    except ImportError as exc:
        raise MergeError(f"A required Excel reader is missing for {path.suffix}: {exc}") from exc
    except Exception as exc:
        raise MergeError(
            f"Could not read {path.name}. It may be damaged, password-protected, or unsupported."
        ) from exc
    frame.columns = [str(column) for column in frame.columns]
    if list(frame.columns) != original_headers:
        frame.columns = original_headers
    validate_columns(list(frame.columns), path.name)
    return frame


def schema_difference(baseline: Sequence[str], actual: Sequence[str]) -> str | None:
    if list(actual) == list(baseline):
        return None
    missing = [column for column in baseline if column not in actual]
    extra = [column for column in actual if column not in baseline]
    details = []
    if len(actual) != len(baseline):
        details.append(f"column count {len(actual)} instead of {len(baseline)}")
    if missing:
        details.append(f"missing {missing}")
    if extra:
        details.append(f"extra {extra}")
    if not missing and not extra:
        details.append("column order differs")
    return "; ".join(details)


def resolve_output_path(
    output_root: os.PathLike[str] | str | None = None,
    output_path: os.PathLike[str] | str | None = None,
) -> Path:
    if output_path:
        resolved = Path(output_path).expanduser().resolve()
        if resolved.suffix.lower() != ".xlsx":
            raise MergeError("The output filename must end with .xlsx.")
        resolved.parent.mkdir(parents=True, exist_ok=True)
        return resolved
    root = Path(output_root).expanduser().resolve() if output_root else Path.cwd().resolve()
    merged_dir = root / "MergedFiles"
    merged_dir.mkdir(parents=True, exist_ok=True)
    return merged_dir / default_output_name()


def validate_destination(output_path: Path, inputs: Sequence[Path], archive_dir: Path | None) -> None:
    input_keys = {os.path.normcase(str(path.resolve())) for path in inputs}
    if os.path.normcase(str(output_path.resolve())) in input_keys:
        raise MergeError("The output file cannot overwrite a source workbook.")
    if not os.access(output_path.parent, os.W_OK):
        raise MergeError(f"The output folder is not writable: {output_path.parent}")
    if archive_dir:
        archive = archive_dir.resolve()
        if os.path.normcase(str(archive)) == os.path.normcase(str(output_path.parent.resolve())):
            raise MergeError("The archive folder cannot be the same as the output folder.")
        for source in inputs:
            if os.path.normcase(str(archive)) == os.path.normcase(str(source.parent.resolve())):
                raise MergeError("The archive folder cannot be the same as a source folder.")


def _canonical_lines(frame: pd.DataFrame) -> Iterable[str]:
    normalized = frame.fillna("")
    separator = "\x1f"
    yield separator.join(str(column) for column in normalized.columns)
    for row in normalized.itertuples(index=False, name=None):
        yield separator.join(str(value) for value in row)


def _checksum(frame: pd.DataFrame) -> str:
    digest = hashlib.sha256()
    for line in _canonical_lines(frame):
        digest.update((line + "\n").encode("utf-8", "surrogatepass"))
    return digest.hexdigest()


def verify_frames(
    merged_path: Path,
    expected: pd.DataFrame,
    diff_dir: Path | None = None,
    max_diffs: int = 1000,
    force_text: bool = True,
) -> VerificationResult:
    try:
        read_options: dict[str, object] = {"engine": "openpyxl", "sheet_name": 0}
        if force_text:
            read_options["dtype"] = str
        merged = pd.read_excel(merged_path, **read_options)
    except Exception as exc:
        raise MergeError("The merged workbook could not be reopened for verification.") from exc
    merged.columns = [str(column) for column in merged.columns]
    expected_columns = [str(column) for column in expected.columns]
    columns_equal = list(merged.columns) == expected_columns
    if not columns_equal:
        return VerificationResult(False, "columns_mismatch", len(merged), len(expected), False)
    if len(merged) != len(expected):
        return VerificationResult(False, "row_count_mismatch", len(merged), len(expected), True)

    merged_normalized = merged.fillna("")
    expected_normalized = expected.fillna("")
    merged_checksum = _checksum(merged_normalized)
    expected_checksum = _checksum(expected_normalized)
    same = merged_normalized.values == expected_normalized.values
    if same.all():
        return VerificationResult(
            True,
            "ok",
            len(merged),
            len(expected),
            True,
            True,
            merged_checksum,
            expected_checksum,
        )

    row_indices, column_indices = np.where(~same)
    column_counts = {
        str(column): int((merged_normalized.iloc[:, index] != expected_normalized.iloc[:, index]).sum())
        for index, column in enumerate(merged_normalized.columns)
    }
    diff_path = None
    if diff_dir:
        diff_dir.mkdir(parents=True, exist_ok=True)
        diff_path = diff_dir / default_output_name("verify-diff").replace(".xlsx", ".csv")
        records = [
            {
                "row_index": int(row),
                "column": str(merged_normalized.columns[column]),
                "expected": expected_normalized.iat[row, column],
                "merged": merged_normalized.iat[row, column],
            }
            for row, column in zip(row_indices[:max_diffs], column_indices[:max_diffs])
        ]
        pd.DataFrame(records).to_csv(diff_path, index=False, encoding="utf-8")
    return VerificationResult(
        False,
        "values_mismatch",
        len(merged),
        len(expected),
        True,
        False,
        merged_checksum,
        expected_checksum,
        diff_path,
        column_counts,
    )


def _atomic_write_excel(frame: pd.DataFrame, output_path: Path, sheet_name: str) -> None:
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent, delete=False
        ) as handle:
            temp_path = Path(handle.name)
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        os.replace(temp_path, output_path)
    except PermissionError as exc:
        raise MergeError("The output workbook is open or the output folder is not writable.") from exc
    except Exception as exc:
        raise MergeError(f"Could not write the merged workbook: {output_path.name}") from exc
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _safe_move(source: Path, destination: Path) -> Path:
    destination.mkdir(parents=True, exist_ok=True)
    target = destination / source.name
    counter = 1
    while target.exists():
        target = destination / f"{source.stem}_copy{counter}{source.suffix}"
        counter += 1
    shutil.move(str(source), str(target))
    return target


def move_sources(paths: Sequence[Path], archive_dir: Path) -> tuple[list[tuple[Path, Path]], list[str]]:
    moved: list[tuple[Path, Path]] = []
    errors: list[str] = []
    for source in paths:
        try:
            moved.append((source, _safe_move(source, archive_dir)))
        except Exception as exc:
            errors.append(f"{source.name}: {exc}")
    return moved, errors


def merge_workbooks(
    input_paths: Sequence[os.PathLike[str] | str],
    *,
    output_root: os.PathLike[str] | str | None = None,
    output_path: os.PathLike[str] | str | None = None,
    sheet_name: str = "Sheet1",
    force_text: bool = True,
    verify: bool = True,
    move_after_success: bool = False,
    archive_dir: os.PathLike[str] | str | None = None,
    progress: ProgressCallback | None = None,
) -> MergeResult:
    inputs = normalize_input_paths(input_paths)
    validate_input_paths(inputs)
    output = resolve_output_path(output_root, output_path)
    archive = Path(archive_dir).expanduser().resolve() if archive_dir else None
    if move_after_success and archive is None:
        archive = output.parent.parent / f"archived-sources-{_stamp()}-{secrets.token_hex(3).upper()}"
    validate_destination(output, inputs, archive if move_after_success else None)
    safe_sheet_name = sanitize_sheet_name(sheet_name)

    total_steps = len(inputs) + 1 + (1 if verify else 0) + (1 if move_after_success else 0)
    completed = 0
    frames: list[pd.DataFrame] = []
    per_file: list[tuple[Path, int]] = []
    baseline: list[str] | None = None

    for path in inputs:
        frame = read_first_sheet(path, force_text)
        columns = list(frame.columns)
        if baseline is None:
            baseline = columns
        else:
            difference = schema_difference(baseline, columns)
            if difference:
                raise SchemaMismatchError(f"Schema mismatch in {path.name}: {difference}.")
        frames.append(frame)
        per_file.append((path, len(frame)))
        completed += 1
        if progress:
            progress(completed, total_steps, f"Read {path.name}: {len(frame):,} rows")

    total_rows = sum(rows for _, rows in per_file)
    if total_rows + 1 > EXCEL_MAX_ROWS:
        raise MergeError(
            f"The merged workbook would contain {total_rows + 1:,} rows including its header, "
            f"which exceeds Excel's {EXCEL_MAX_ROWS:,}-row limit."
        )

    merged = pd.concat(frames, axis=0, ignore_index=True)
    _atomic_write_excel(merged, output, safe_sheet_name)
    completed += 1
    if progress:
        progress(completed, total_steps, f"Wrote {output.name}")

    verification = None
    if verify:
        verification = verify_frames(
            output,
            merged,
            output.parent.parent / "Verification",
            force_text=force_text,
        )
        completed += 1
        if progress:
            progress(completed, total_steps, "Verification passed" if verification.passed else "Verification failed")

    moved: list[tuple[Path, Path]] = []
    move_errors: list[str] = []
    if move_after_success:
        if verification is not None and not verification.passed:
            raise MergeError("Verification failed, so source files were not moved.")
        assert archive is not None
        moved, move_errors = move_sources(inputs, archive)
        completed += 1
        if progress:
            progress(completed, total_steps, f"Moved {len(moved)} of {len(inputs)} source files")

    return MergeResult(
        output,
        total_rows,
        tuple(per_file),
        verification,
        tuple(moved),
        tuple(move_errors),
    )


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append the first worksheet from multiple Excel files using a strict schema."
    )
    parser.add_argument("files", nargs="*", help="Input Excel files in append order")
    parser.add_argument("--pattern", action="append", default=[], help="Glob pattern for additional files")
    parser.add_argument("--pattern-sort", choices=["name", "none"], default="name")
    parser.add_argument("--outdir", help="Root folder for automatically named output")
    parser.add_argument("--output", help="Exact .xlsx output path")
    parser.add_argument("--sheet-out", default="Sheet1", help="Output worksheet name")
    parser.add_argument("--no-force-text", action="store_false", dest="force_text")
    parser.add_argument("--no-verify", action="store_false", dest="verify")
    parser.add_argument("--move-sources", action="store_true")
    parser.add_argument("--move-dest", help="Archive folder used with --move-sources")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--gui", action="store_true", help="Launch the graphical interface")
    parser.set_defaults(force_text=True, verify=True)
    return parser.parse_args(argv)


class MergeGUI:
    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.ttk = ttk
        self.root = tk.Tk()
        self.root.title("Excel Merge Utility")
        self.root.geometry("1060x720")
        self.root.minsize(900, 620)
        self.root.configure(bg="#0f172a")

        self.files: list[Path] = []
        self.last_folder = Path.home()
        self.outdir = tk.StringVar(value="")
        self.archive_dir = tk.StringVar(value="")
        self.verify = tk.BooleanVar(value=True)
        self.move_sources_var = tk.BooleanVar(value=False)
        self.verbose = tk.BooleanVar(value=False)
        self.status_text = tk.StringVar(value="Select Excel files to begin.")
        self.count_text = tk.StringVar(value="0 files selected")
        self.events: queue.Queue[tuple[str, object]] = queue.Queue()
        self.controls: list[object] = []
        self.is_merging = False
        self.last_output: Path | None = None

        self._build_ui()
        self.root.after(100, self._drain_events)

    def _button(self, parent, text: str, command, *, primary: bool = False):
        color = "#22c55e" if primary else "#1f2937"
        button = self.tk.Button(
            parent,
            text=text,
            command=command,
            bg=color,
            fg="#050a12" if primary else "white",
            activebackground="#16a34a" if primary else "#334155",
            activeforeground="white",
            bd=0,
            padx=11,
            pady=7,
            cursor="hand2",
        )
        self.controls.append(button)
        return button

    def _build_ui(self) -> None:
        tk = self.tk
        header = tk.Frame(self.root, bg="#0891b2", height=100)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Excel Merge Utility", bg="#0891b2", fg="white", font=("Segoe UI", 22, "bold")).pack(
            anchor="w", padx=24, pady=(18, 0)
        )
        tk.Label(
            header,
            text="Select files, confirm their order, choose an output folder, and merge safely.",
            bg="#0891b2",
            fg="#ecfeff",
            font=("Segoe UI", 10),
        ).pack(anchor="w", padx=24, pady=(2, 0))

        body = tk.Frame(self.root, bg="#0f172a")
        body.pack(fill="both", expand=True, padx=18, pady=16)
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        left = tk.Frame(body, bg="#0f172a")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        actions = tk.Frame(left, bg="#0f172a")
        actions.pack(fill="x", pady=(0, 8))
        self._button(actions, "Select Files", self._pick).pack(side="left", padx=(0, 6))
        self._button(actions, "Remove Selected", self._remove_selected).pack(side="left", padx=6)
        self._button(actions, "Move Up", self._move_up).pack(side="left", padx=6)
        self._button(actions, "Move Down", self._move_down).pack(side="left", padx=6)
        self._button(actions, "Clear", self._clear).pack(side="left", padx=6)

        tk.Label(left, textvariable=self.count_text, bg="#0f172a", fg="#cbd5e1").pack(anchor="w", pady=(0, 4))
        list_frame = tk.Frame(left, bg="#0f172a")
        list_frame.pack(fill="both", expand=True)
        self.listbox = tk.Listbox(
            list_frame,
            selectmode="extended",
            bg="#111827",
            fg="white",
            selectbackground="#0369a1",
            font=("Consolas", 10),
            highlightthickness=1,
            highlightbackground="#334155",
        )
        self.listbox.pack(side="left", fill="both", expand=True)
        scroll = tk.Scrollbar(list_frame, command=self.listbox.yview)
        scroll.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=scroll.set)

        tk.Label(left, text="Activity", bg="#0f172a", fg="#cbd5e1", font=("Segoe UI", 10, "bold")).pack(
            anchor="w", pady=(10, 4)
        )
        self.log_text = tk.Text(
            left, height=10, state="disabled", bg="#0b1220", fg="#e5e7eb", wrap="word", font=("Consolas", 9)
        )
        self.log_text.pack(fill="x")

        right = tk.Frame(body, bg="#111827", highlightthickness=1, highlightbackground="#334155")
        right.grid(row=0, column=1, sticky="nsew")
        tk.Label(right, text="Output folder", bg="#111827", fg="white", font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=14, pady=(16, 4)
        )
        output_row = tk.Frame(right, bg="#111827")
        output_row.pack(fill="x", padx=14)
        self.output_entry = tk.Entry(output_row, textvariable=self.outdir, state="readonly")
        self.output_entry.pack(side="left", fill="x", expand=True)
        self._button(output_row, "Browse…", self._browse_output).pack(side="left", padx=(8, 0))

        self.verify_box = tk.Checkbutton(
            right,
            text="Verify every merged value",
            variable=self.verify,
            bg="#111827",
            fg="white",
            activebackground="#111827",
            selectcolor="#111827",
        )
        self.verify_box.pack(anchor="w", padx=14, pady=(16, 4))
        self.controls.append(self.verify_box)
        self.verbose_box = tk.Checkbutton(
            right,
            text="Print technical errors to the console",
            variable=self.verbose,
            bg="#111827",
            fg="white",
            activebackground="#111827",
            selectcolor="#111827",
        )
        self.verbose_box.pack(anchor="w", padx=14, pady=4)
        self.controls.append(self.verbose_box)
        self.move_box = tk.Checkbutton(
            right,
            text="Move source files after success",
            variable=self.move_sources_var,
            command=self._toggle_archive,
            bg="#111827",
            fg="white",
            activebackground="#111827",
            selectcolor="#111827",
        )
        self.move_box.pack(anchor="w", padx=14, pady=4)
        self.controls.append(self.move_box)

        tk.Label(right, text="Archive folder (optional)", bg="#111827", fg="#cbd5e1").pack(
            anchor="w", padx=14, pady=(8, 4)
        )
        archive_row = tk.Frame(right, bg="#111827")
        archive_row.pack(fill="x", padx=14)
        self.archive_entry = tk.Entry(archive_row, textvariable=self.archive_dir, state="disabled")
        self.archive_entry.pack(side="left", fill="x", expand=True)
        self.archive_button = self._button(archive_row, "Browse…", self._browse_archive)
        self.archive_button.pack(side="left", padx=(8, 0))
        self.archive_button.configure(state="disabled")

        self.progress = self.ttk.Progressbar(right, mode="determinate", maximum=1)
        self.progress.pack(fill="x", padx=14, pady=(24, 6))
        tk.Label(right, textvariable=self.status_text, bg="#111827", fg="#cbd5e1", wraplength=330, justify="left").pack(
            anchor="w", padx=14
        )
        self._button(right, "Merge Now", self._start_merge, primary=True).pack(fill="x", padx=14, pady=(18, 8))
        self.open_button = self._button(right, "Open Output Folder", self._open_output)
        self.open_button.pack(fill="x", padx=14, pady=(0, 16))
        self.open_button.configure(state="disabled")

    def _refresh_list(self) -> None:
        self.listbox.delete(0, "end")
        for path in self.files:
            self.listbox.insert("end", str(path))
        self.count_text.set(f"{len(self.files)} file{'s' if len(self.files) != 1 else ''} selected")

    def _pick(self) -> None:
        selected = self.filedialog.askopenfilenames(
            title="Select Excel files in append order",
            initialdir=str(self.last_folder),
            filetypes=[("Excel workbooks", "*.xlsx *.xlsm *.xltx *.xltm *.xls")],
        )
        if not selected:
            return
        self.last_folder = Path(selected[0]).parent
        combined = normalize_input_paths([*self.files, *selected])
        added = len(combined) - len(self.files)
        self.files = combined
        self._refresh_list()
        self.status_text.set(f"Added {added} file(s). Duplicate selections were ignored.")

    def _selected_indices(self) -> list[int]:
        return list(self.listbox.curselection())

    def _remove_selected(self) -> None:
        selected = set(self._selected_indices())
        self.files = [path for index, path in enumerate(self.files) if index not in selected]
        self._refresh_list()

    def _clear(self) -> None:
        self.files.clear()
        self._refresh_list()
        self.status_text.set("File list cleared.")

    def _move_up(self) -> None:
        selected = self._selected_indices()
        if not selected:
            return
        for index in selected:
            if index > 0 and index - 1 not in selected:
                self.files[index - 1], self.files[index] = self.files[index], self.files[index - 1]
        self._refresh_list()
        for index in [max(0, item - 1) for item in selected]:
            self.listbox.select_set(index)

    def _move_down(self) -> None:
        selected = self._selected_indices()
        if not selected:
            return
        for index in reversed(selected):
            if index < len(self.files) - 1 and index + 1 not in selected:
                self.files[index + 1], self.files[index] = self.files[index], self.files[index + 1]
        self._refresh_list()
        for index in [min(len(self.files) - 1, item + 1) for item in selected]:
            self.listbox.select_set(index)

    def _browse_output(self) -> None:
        selected = self.filedialog.askdirectory(title="Choose output folder", initialdir=str(self.last_folder))
        if selected:
            self.outdir.set(selected)
            self.last_folder = Path(selected)

    def _browse_archive(self) -> None:
        selected = self.filedialog.askdirectory(title="Choose archive folder", initialdir=str(self.last_folder))
        if selected:
            self.archive_dir.set(selected)
            self.last_folder = Path(selected)

    def _toggle_archive(self) -> None:
        state = "normal" if self.move_sources_var.get() and not self.is_merging else "disabled"
        self.archive_entry.configure(state=state)
        self.archive_button.configure(state=state)

    def _set_controls(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        for control in self.controls:
            try:
                control.configure(state=state)
            except Exception:
                pass
        self.output_entry.configure(state="readonly" if enabled else "disabled")
        self.open_button.configure(state="normal" if enabled and self.last_output else "disabled")
        if enabled:
            self._toggle_archive()

    def _log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{datetime.now():%H:%M:%S}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _start_merge(self) -> None:
        if self.is_merging:
            return
        if not self.files:
            self.messagebox.showerror("No files selected", "Select one or more Excel files.")
            return
        if not self.outdir.get().strip():
            self.messagebox.showerror("Output folder required", "Choose an output folder before merging.")
            return
        options = {
            "input_paths": tuple(self.files),
            "output_root": self.outdir.get().strip(),
            "verify": self.verify.get(),
            "move_after_success": self.move_sources_var.get(),
            "archive_dir": self.archive_dir.get().strip() or None,
        }
        self.is_merging = True
        self.last_output = None
        self._set_controls(False)
        self.progress.configure(maximum=max(1, len(self.files) + 3), value=0)
        self.status_text.set("Merging…")
        self._log("Starting merge.")
        verbose_console = bool(self.verbose.get())
        threading.Thread(target=self._worker, args=(options, verbose_console), daemon=True).start()

    def _worker(self, options: dict[str, object], verbose_console: bool) -> None:
        def progress(completed: int, total: int, message: str) -> None:
            self.events.put(("progress", (completed, total, message)))

        try:
            result = merge_workbooks(**options, progress=progress)
            self.events.put(("success", result))
        except Exception as exc:
            if verbose_console:
                traceback.print_exc()
            self.events.put(("error", str(exc)))

    def _drain_events(self) -> None:
        try:
            while True:
                event, payload = self.events.get_nowait()
                if event == "progress":
                    completed, total, message = payload
                    self.progress.configure(maximum=total, value=completed)
                    self.status_text.set(message)
                    self._log(message)
                elif event == "success":
                    self._finish_success(payload)
                elif event == "error":
                    self._finish_error(str(payload))
        except queue.Empty:
            pass
        self.root.after(100, self._drain_events)

    def _finish_success(self, result: MergeResult) -> None:
        self.is_merging = False
        self.last_output = result.output_path
        self._set_controls(True)
        warning = ""
        if result.verification and not result.verification.passed:
            warning = "\nVerification did not pass. Source files were not moved."
        if result.move_errors:
            warning += f"\n{len(result.move_errors)} source file(s) could not be archived."
        self.status_text.set(f"Completed: {result.total_rows:,} rows")
        self._log(f"Output: {result.output_path}")
        self._save_log()
        self.messagebox.showinfo(
            "Merge complete",
            f"Merged {result.total_rows:,} rows.\n\nOutput:\n{result.output_path}{warning}",
        )

    def _finish_error(self, message: str) -> None:
        self.is_merging = False
        self._set_controls(True)
        self.status_text.set("Merge failed. Review the message and try again.")
        self._log(f"ERROR: {message}")
        self._save_log()
        self.messagebox.showerror("Merge failed", message)

    def _save_log(self) -> None:
        output_root = self.outdir.get().strip()
        if not output_root:
            return
        try:
            log_dir = Path(output_root).expanduser().resolve() / "Log"
            log_dir.mkdir(parents=True, exist_ok=True)
            log_path = log_dir / default_output_name("merge-log").replace(".xlsx", ".txt")
            log_path.write_text(self.log_text.get("1.0", "end").strip() + "\n", encoding="utf-8")
        except Exception as exc:
            self._log(f"WARNING: Could not save the activity log: {exc}")

    def _open_output(self) -> None:
        if self.last_output and self.last_output.parent.exists():
            if os.name == "nt":
                os.startfile(self.last_output.parent)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(self.last_output.parent)])
            else:
                subprocess.Popen(["xdg-open", str(self.last_output.parent)])

    def run(self) -> None:
        self.root.mainloop()


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.gui:
        MergeGUI().run()
        return 0

    inputs = [*args.files, *expand_patterns(args.pattern, args.pattern_sort)]
    if not inputs:
        print("ERROR: Provide Excel files, a --pattern, or use --gui.", file=sys.stderr)
        return 2

    def report(completed: int, total: int, message: str) -> None:
        if args.verbose:
            print(f"[{completed}/{total}] {message}")

    try:
        result = merge_workbooks(
            inputs,
            output_root=args.outdir,
            output_path=args.output,
            sheet_name=args.sheet_out,
            force_text=args.force_text,
            verify=args.verify,
            move_after_success=args.move_sources,
            archive_dir=args.move_dest,
            progress=report,
        )
    except SchemaMismatchError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 3
    except MergeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    print(f"Merge complete: {result.total_rows:,} rows")
    print(f"Output: {result.output_path}")
    if result.verification:
        print(f"Verification: {'PASSED' if result.verification.passed else 'FAILED'}")
        if not result.verification.passed:
            return 5
    if result.move_errors:
        print("Archive warnings:", file=sys.stderr)
        for error in result.move_errors:
            print(f" - {error}", file=sys.stderr)
        return 4
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
